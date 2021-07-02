import sqlite3
import tkinter as tk
from tkinter import ttk
from tkinter import *
from PIL import Image,ImageTk
import sqlite3
import datetime
import webbrowser
import subprocess
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen.canvas import *
from reportlab.lib.utils import ImageReader
import pytz
import openpyxl
from openpyxl import *
import xlwt 
from xlwt import Workbook
from tkinter import messagebox
from functools import partial
import numpy as np
import os, sys
import pandas as pd


root=tk.Tk()
root.geometry("%dx%d+0+0"%(root.winfo_screenwidth(),root.winfo_screenheight()))
ancho=root.winfo_screenwidth()
alto=root.winfo_screenheight()
uan= ancho/13.66
ual=alto/7.68
root.destroy()
class Window(tk.Tk):
    def __init__(self,*args, **kwargs):
        tk.Tk.__init__(self,*args,**kwargs)
        self.geometry(str(ancho)+'x'+str(alto))
        self.title('Carta Electrónica')
        self.frame=None
        self.precios= {"Hamburguesa Sencilla":3000,"Hamburguesa Especial":7000,"Papa Frita":3000,"Yuca Frita":3000,
                       "Queso":3000,"Salchipapa Sencilla":3000,"Salchipapa Especial":6000,"Salvajada":20000,"Perro Sencillo":3000,
                       "Perro Especial":6000,"Coca Cola":4000,"Pepsi":3000,"Aguila":4500}
        
        self.productos={"Hamburguesa Sencilla":0,"Hamburguesa Especial":0,"Papa Frita":0,"Yuca Frita":0,
                       "Queso":0,"Salchipapa Sencilla":0,"Salchipapa Especial":0,"Salvajada":0,"Perro Sencillo":0,
                       "Perro Especial":0,"Coca Cola":0,"Pepsi":0,"Aguila":0}
        self.pedido=int(np.loadtxt("cuenta_pedido.txt"))
        self.valor=0
        self.valor= sum([self.productos[producto]*self.precios[producto] for producto in self.productos])
        self.usuario=""
        self.telefono_usuario=""
        self.cedula_usuario=""
        
        self.show_frame(Login)
        
    def default(self):
        self.productos={"Hamburguesa Sencilla":0,"Hamburguesa Especial":0,"Papa Frita":0,"Yuca Frita":0,
                       "Queso":0,"Salchipapa Sencilla":0,"Salchipapa Especial":0,"Salvajada":0,"Perro Sencillo":0,
                       "Perro Especial":0,"Coca Cola":0,"Pepsi":0,"Aguila":0}
        self.valor=0
        self.usuario=""
        self.telefono_usuario=""
        self.cedula_usuario=""
        self.show_frame(Salchi)
        
    def actualizar_valor(self):
        self.valor=sum([self.productos[producto]*self.precios[producto] for producto in self.productos])
    
    def show_frame(self, frame_class):
        new_frame = frame_class(self)
        if self.frame != None:
            self.frame.pack_forget()
        self.frame = new_frame
        self.frame.pack(side = 'right', expand = True, fill= 'both', ipadx = 210)
        self.frame.tkraise()
        
class Opciones(tk.Frame):
    def __init__(self,parent):
        
        tk.Frame.__init__(self,parent)
        
        
        def reporte(self,parent):
            
            precios_mod= {"H. Sencilla":3000,"H. Especial":7000,"Papa Frita":3000,"Yuca Frita":3000,
                       "Queso":3000,"S. Sencilla":3000,"S. Especial":6000,"Salvajada":20000,"P. Sencillo":3000,
                       "P. Especial":6000,"Coca Cola":4000,"Pepsi":3000,"Aguila":4500}
            
            import matplotlib.pylab as plt
            
            precios_mod= {"H. Sencilla":3000,"H. Especial":7000,"Papa Frita":3000,"Yuca Frita":3000,
                       "Queso":3000,"S. Sencilla":3000,"S. Especial":6000,"Salvajada":20000,"P. Sencillo":3000,
                       "P. Especial":6000,"Coca Cola":4000,"Pepsi":3000,"Aguila":4500}
            precios= {"Hamburguesa Sencilla":3000,"Hamburguesa Especial":7000,"Papa Frita":3000,"Yuca Frita":3000,
                                   "Queso":3000,"Salchipapa Sencilla":3000,"Salchipapa Especial":6000,"Salvajada":20000,"Perro Sencillo":3000,
                                   "Perro Especial":6000,"Coca Cola":4000,"Pepsi":3000,"Aguila":4500}
            import matplotlib.pylab as plt


            plt.figure(figsize=(15,8))
            plt.title("Precios de los productos",size=20)
            plt.ylabel("Valor en COP ($)",size=15)
            plt.bar(precios_mod.keys(),precios_mod.values(),color="green")
            plt.grid() 
            plt.savefig("precios.png")
            plt.show()

            ex=pd.read_excel("registro.xlsx")
            veces={}

            for i in precios:
                lista=[]
                for j in ex[i]:
                    if str(j)=="nan":
                        lista.append(0)
                    else:
                        lista.append(j)

                veces[i]= sum(lista)

            plt.figure(figsize=(15,8))
            plt.title("Veces que se han comprado los productos",size=20)
            plt.ylabel("Cantidad de Veces (#)",size=15)
            plt.grid()  
            plt.bar(precios_mod.keys(),veces.values())
            plt.savefig("frecuencia_productos.png")
            plt.show()

            def mas_comprado_(veces):
                mayor=0
                mas= ""
                lista=[]
                contador=0
                for i in veces.keys():
                    if veces[i]>mayor:
                        mayor=veces[i]
                        mas=i


                for i in veces.keys():
                    if veces[i]==mayor:
                        contador+=1
                        lista.append(i)

                retorno=0

                if contador==1:

                    retorno= mas
                else:
                    retorno=lista

                return retorno, mayor

            def mas_comprado(veces):
                mayor=0
                mas= ""
                for i in veces.keys():
                    if veces[i]>mayor:
                        mayor=veces[i]
                        mas=i

                return mas, mayor


            def menos_comprado_(veces):
                menor=10**9
                menos= ""
                lista=[]
                contador=0
                for i in veces.keys():
                    if veces[i]<menor:
                        menor=veces[i]
                        menos=i

                for i in veces.keys():
                    if veces[i]==menor:
                        contador+=1
                        lista.append(i)

                retorno=0

                if contador==1:

                    retorno= menos
                else:
                    retorno=lista

                return retorno, menor

            def menos_comprado(veces):
                menor=10**9
                menos= ""
                for i in veces.keys():
                    if veces[i]<menor:
                        menor=veces[i]
                        menos=i

                return menos, menor

            mayor=0
            nombre=""
            cedula=""
            vec={}
            indices={}
            for i in range(len(ex["Cédula"])):

                if ex["Cédula"][i] not in vec:
                    vec[ex["Cédula"][i]]=1
                    indices[ex["Cédula"][i]]=[i]
                else:
                    vec[ex["Cédula"][i]]+=1
                    indices[ex["Cédula"][i]].append(i)
            for i in vec:
                if vec[i]>mayor and i!="No puso cedula":
                    mayor=vec[i]
                    cedula=i
                    nombre=ex["Nombre"][indices[i][0]]
                    tel=ex["Teléfono"][indices[i][0]]

            y=30
            pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
            c = canvas.Canvas("estadistico.pdf",pagesize=(650,1000))
            c.setFont('Arial',30)
            text = "Reporte de Estadística"
            text_width = c.stringWidth(text,'Arial',30)
            c.drawString((650-text_width)/2, 970, text)
            #
            c.setFont('Arial',25)
            text = "Precios"
            text_width = c.stringWidth(text,'Arial',25)
            c.drawString((650-text_width)/2, 910, text)

            img1= ImageReader("precios.png")
            c.drawImage(img1, 10,500, 600,400,mask='auto')

            text = "Frecuencia de Compra"
            text_width = c.stringWidth(text,'Arial',25)
            c.drawString((650-text_width)/2, 480, text)

            img1= ImageReader("frecuencia_productos.png")
            c.drawImage(img1, 10,50, 600,400,mask='auto')

            c.showPage()

            c.setFont('Arial',25) 
            text = "Artículo Más y menos Comprado"
            text_width = c.stringWidth(text,'Arial',25)
            c.drawString((650-text_width)/2, 960, text)
            c.setFont('Arial',20) 
            mas_com= mas_comprado(veces)[0]
            text= "El artículo más comprado es: " + mas_com + ". ("+str(int(mas_comprado(veces)[1]))+" veces)"
            c.drawString(20, 900, text)

            menos_com= menos_comprado(veces)[0]
            text= "El artículo menos comprado es: " + menos_com + ". ("+str(int(menos_comprado(veces)[1]))+" veces)"
            c.drawString(20, 840, text)

            c.setFont('Arial',25) 
            text = "Mejor Cliente"
            text_width = c.stringWidth(text,'Arial',25)
            c.drawString((650-text_width)/2, 760, text)
            c.setFont('Arial',20) 

            text =  nombre+" CC. "+cedula 
            text_width = c.stringWidth(text,'Arial',20)
            c.drawString((650-text_width)/2, 700, text)

            c.setFont('Arial',25) 
            text = "Ingresos Totales"
            text_width = c.stringWidth(text,'Arial',25)
            c.drawString((650-text_width)/2, 620, text)

            text= str(sum([veces[i]*precios[i] for i in veces])) + " COP ($)"
            text_width = c.stringWidth(text,'Arial',25)
            c.drawString((650-text_width)/2, 550, text)

            c.save()
            chrome_path= "C:\Program Files\Google\Chrome\Application\chrome.exe"
            subprocess.Popen([chrome_path, "estadistico.pdf"])

            
        
        background=Image.open("background_vinotinto.jpg")
        background=background.resize((ancho,alto),Image.ANTIALIAS)
        bk_photo = ImageTk.PhotoImage(background,master=self)
        
        bk_label=tk.Label(self,image=bk_photo)
        bk_label.image=bk_photo
        bk_label.place(x=0, y=0)
        
        bsal_label=tk.Label(self,text="¿A dónde quiere ir ?",font=("Brush Script MT",40))
        bsal_label.place(x=ancho/2.65, y=(0.3*alto/3.3))
        
        
        bsal_label=tk.Label(self,text="Reporte de Estadística",font=("Brush Script MT",30))
        bsal_label.place(x=(0.95*ancho/5), y=1.5*alto/5)
        
        imagen_bsal=Image.open("reporte_estadistico.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(ancho/5)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_rep= Button(self,text="",image=photo_bsal,command=lambda: reporte(self,parent))
        button_rep.image=photo_bsal
        button_rep.place(x=(ancho/5),y=2*alto/5)
        
        bsal_label=tk.Label(self,text="Software",font=("Brush Script MT",30))
        bsal_label.place(x=(3.27*ancho/5), y=1.5*alto/5)
        
        imagen_bsal=Image.open("software.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(ancho/5)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_= Button(self,text="",image=photo_bsal,command=lambda: parent.show_frame(Salchi))
        button_.image=photo_bsal
        button_.place(x=(3*ancho/5),y=2*alto/5)
        
        
class Login(tk.Frame):
    def __init__(self,parent):
        tk.Frame.__init__(self,parent)
        background=Image.open("background_vinotinto.jpg")
        background=background.resize((300,450),Image.ANTIALIAS)
        bk_photo = ImageTk.PhotoImage(background,master=self)
        
        bk_label=tk.Label(self,image=bk_photo)
        bk_label.image=bk_photo
        bk_label.place(x=(ancho-300)/2, y=(alto-450)/2)
        
        animado=Image.open("animado.jpg")
        animado=animado.resize((200,350),Image.ANTIALIAS)
        an_photo = ImageTk.PhotoImage(animado,master=self)
        
        bk_label1=tk.Label(self,image=an_photo)
        bk_label1.image=an_photo
        bk_label1.place(x=(ancho-(2*uan))/2, y=(alto-(3.5*ual))/2)
        
        Label(self,text = "Software de \n Carta Electrónica ",font=("Helvetica", 18),bg="green").place(x=(ancho-(2*uan))/2,
                                                                                                       y=(2.5*ual))
        
        usernameLabel = Label(self, text="Usuario",font=("Helvetica", 12),bg="white").place(x=(ancho-(2.5*uan))/2,y=(alto/2)+
                                                                                            (0.8*ual))
        self.usuario = Entry(self,font=("Helvetica", 12))
        self.usuario.place(x=(ancho-(1.25*uan))/2,y=(alto/2)+(0.81*ual))
        
        passwordLabel = Label(self, text=" Clave  ",font=("Helvetica", 12),bg="white").place(x=(ancho-(2.5*uan))/2,y=(alto/2)+
                                                                                             (1.1*ual))
        self.clave = Entry(self,show='*',font=("Helvetica", 12))
        self.clave.place(x=(ancho-(1.25*uan))/2,y=(alto/2)+(1.11*ual))
        
        def validateLogin(self,username, password):
            if self.usuario.get()== username and self.clave.get()==password:
                parent.show_frame(Opciones)
            else:
                messagebox.showerror(title="Error", 
                                    message="El nombre y/o la contraseña son incorrectas")
                
        loginButton = Button(self, text="Iniciar Sesión",font=("Helvetica", 12), command=lambda: validateLogin(self,"admin","hola")).place(x=(ancho-105)/2,y=(alto/2)+140)
        
class Salchi(tk.Frame):
    def __init__(self,parent):
        #BACKGROUND#
        tk.Frame.__init__(self,parent)
        def serial(numero):
            s= "000000"
            resultado= s[0:len(s)-len(str(numero))] + str(numero)
            return resultado
        def hacer_pedido(self, parent):
            nombre=parent.usuario
            tel= parent.telefono_usuario
            cedula= parent.cedula_usuario
            if len(tel)==0:
                tel= "Null"
            if len(nombre)==0:
                nombre="No puso nombre"
            if len(cedula)==0:
                cedula="No puso cedula"
            
            lista_p=[]
            productos_selec={}
            indices= {}
            contador=0
            cont=0
            for i in parent.productos:
                if parent.productos[i]>0:
                    contador+=1
                    productos_selec[i]=parent.productos[i]
                cont+=1
                indices[i]=cont                 
            if contador>0:
                parent.pedido+=1
                y=30
                pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
                c = canvas.Canvas("invoice.pdf",pagesize=(300,270+(len(productos_selec)*40)),bottomup=0)
                c.setFont('Arial',15)
                text = "Recibo Electrónico"
                text_width = c.stringWidth(text,'Arial',15)
                c.drawString((300-text_width)/2, y, text)
                #Separador
                y=y+20
                text="------------------------------------------------------------"
                c.drawString(0, y, text)
                #Info Negocio
                c.setFont('Arial',9)
                y=y+20
                text="Dirección: Diagonal 21 No. 29-56, Barrio Sabanas del Valle"
                c.drawString(20, y, text)
                y=y+20
                text="Telefono: 312 5698741"
                c.drawString(20, y, text)
                #Separador
                c.setFont('Arial',15)
                y=y+20
                text="------------------------------------------------------------"
                c.drawString(0, y, text)
                #Separador
                c.setFont('Arial',9)
                y=y+20
                text="Fecha: "+ datetime.datetime.now(pytz.timezone('America/Bogota')).strftime("%m/%d/%Y, %H:%M:%S")
                c.drawString(20, y, text)
                c.setFont('Arial',9)
                y=y+20
                text="Nombre de Cliente: "+nombre
                c.drawString(20, y, text)
                y=y+20
                text="Cédula del Cliente: "+cedula
                c.drawString(20, y, text)
                y=y+20
                text="Teléfono de Cliente: "+tel
                c.drawString(20, y, text)
                y=y+20
                text="ID del Pedido: "+ serial(parent.pedido)
                c.drawString(20, y, text)
                i=0
                c.setFont('Arial',8)
                for j in productos_selec:
                    y=y+20
                    text= "["+str(i+1)+"]"+" "+j+" x "+str(productos_selec[j])
                    c.drawString(20, y, text)
                    y=y+20
                    text= "Precio: "+str(parent.precios[j]*productos_selec[j]) + " $"
                    c.drawString(20, y, text)
                    i+=1
                #Separador
                c.setFont('Arial',15)
                y=y+20
                text="------------------------------------------------------------"
                c.drawString(0, y, text)
                #Total a pagar
                # Saving the PDF
                c.setFont('Arial',11)
                y=y+20
                text="Total A Pagar: "+str(sum([productos_selec[i]*parent.precios[i] for i in productos_selec])) + " $"
                c.drawString(20, y, text)
                c.showPage()
                c.save()
                chrome_path= "C:\Program Files\Google\Chrome\Application\chrome.exe"
                subprocess.Popen([chrome_path, "invoice.pdf"])
                # Workbook is created 
                wb=load_workbook("registro.xlsx")
                ws=wb["Hoja de pedidos"] 
                #Valores
                ws.cell(parent.pedido+1, 1).value=serial(parent.pedido)
                lista_p.append(serial(parent.pedido))
                ws.cell(parent.pedido+1, 2).value= datetime.datetime.now(pytz.timezone('America/Bogota')).strftime("%m/%d/%Y, %H:%M:%S")
                lista_p.append(datetime.datetime.now(pytz.timezone('America/Bogota')).strftime("%m/%d/%Y, %H:%M:%S"))
                ws.cell(parent.pedido+1, 3).value=nombre
                lista_p.append(nombre)
                ws.cell(parent.pedido+1, 4).value=cedula
                lista_p.append(cedula)
                ws.cell(parent.pedido+1, 5).value=tel
                lista_p.append(tel)
                
                for i in productos_selec:
                    ws.cell(parent.pedido+1, 5+ indices[i]).value=productos_selec[i]
                
                for j in parent.productos:
                    lista_p.append(parent.productos[j])
                    
                parent.default()   
                ws.cell(parent.pedido+1, 19).value=sum([productos_selec[i]*parent.precios[i] for i in productos_selec])
                lista_p.append(sum([productos_selec[i]*parent.precios[i] for i in productos_selec]))
                wb.save('registro.xlsx') 
                
                def run_query(consulta,datos):
                    with sqlite3.connect("pedidos.db") as conn:
                        cursor=conn.cursor()
                        resultado=cursor.execute(consulta,datos)
                        conn.commit()
                    return resultado
    
                def add_product(tup):
                    query="INSERT INTO Pedidos VALUES( ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
                    parameters=tup
                    run_query(query,parameters)
                    
                add_product(tuple(lista_p))
                
                def cuenta_pedido(parent):
                    mi_archivo = os.open("cuenta_pedido.txt",os.O_RDWR)
                    os.write(mi_archivo,str(parent.pedido).encode())
                    os.close(mi_archivo)
                    
                cuenta_pedido(parent)
            else:
                messagebox.showwarning(title="No se puede Hacer el pedido", 
                                               message="No ha seleccionado ningún elemento")
            
        background=Image.open("background_vinotinto.jpg")
        background=background.resize((ancho,alto),Image.ANTIALIAS)
        bk_photo = ImageTk.PhotoImage(background,master=self)
        
        bk_label=tk.Label(self,image=bk_photo)
        bk_label.image=bk_photo
        bk_label.place(x=0, y=0)
        
        top_banner=Image.open("BANNER.jpeg")
        top_banner=top_banner.resize((int(ancho/1.5),int(alto/5)),Image.ANTIALIAS)
        tb_photo = ImageTk.PhotoImage(top_banner,master=self)
        
        tb_image_label= tk.Label(self,image=tb_photo)
        tb_image_label.image=tb_photo
        tb_image_label.place(x=ancho/6,y=0)
        
        Label(self,text = "Escriba su Nombre",font=("Helvetica", 18)).place(x=3*ancho/15,y=alto/4.47)
        self.usuario=Entry(self,font=("Arial", 16))
        self.usuario.place(x=2.5*ancho/15,y=1.25*alto/4.5,width=3*uan,height=(0.35*ual))
        self.usuario.insert(0,parent.usuario)
        
        Label(self,text = "Escriba su Cédula",font=("Helvetica", 18)).place(x=6.5*ancho/15,y=alto/4.47)
        self.cedula=Entry(self,font=("Arial", 16))
        self.cedula.place(x=(6*ancho/15),y=1.25*alto/4.5,width=3*uan,height=(0.35*ual))
        self.cedula.insert(0,parent.cedula_usuario)
        
        Label(self,text = "Escriba su número ",font=("Helvetica", 18)).place(x=10*ancho/15,y=alto/4.47)
        self.telusuario=Entry(self,font=("Arial", 16))
        self.telusuario.place(x=9.5*ancho/15,y=1.25*alto/4.5,width=3*uan,height=(0.35*ual))
        self.telusuario.insert(0,parent.telefono_usuario)
        
        def actualizar_usuario(self,parent):
            
            if len(str(self.usuario.get()))>0:
            
                if self.usuario.get().replace(" ","").isalpha():
                    parent.usuario=self.usuario.get()

                else:
                    messagebox.showerror(title="Error", 
                                        message="El nombre no puede tener números")
            
            
            if len(str(self.telusuario.get()))>0:
                if self.telusuario.get().isdigit():
                    if len(str(self.telusuario.get()))==10:
                        parent.telefono_usuario=self.telusuario.get()
                    else:
                        messagebox.showerror(title="Error", 
                                        message="El número debe tener 10 dígitos")

                else:
                    messagebox.showerror(title="Error", 
                                        message="El número no puede tener letras")
                    
            if len(str(self.cedula.get()))>0:
                
                if self.cedula.get().isdigit():
                    if len(str(self.cedula.get()))==10:
                        parent.cedula_usuario=self.cedula.get()
                    else:
                        messagebox.showerror(title="Error", 
                                        message="La cédula debe tener 10 dígitos")
                else:
                    messagebox.showerror(title="Error", 
                                        message="La cédula no puede tener letras")
                
                
           
        #Boton selec_usario
        imagen_bsal=Image.open("check.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/20),int(alto/25)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_check= Button(self,text="",image=photo_bsal,command=lambda: actualizar_usuario(self,parent))
        button_check.image=photo_bsal
        button_check.place(x=(13*ancho/15),y=1.25*alto/4.5)
        #Boton salchipapa
        imagen_bsal=Image.open("boton_salchipapa.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_salchi= Button(self,text="",image=photo_bsal,command=lambda: parent.show_frame(Salchipapa))
        button_salchi.image=photo_bsal
        button_salchi.place(x=0,y=1.1*alto/2)
        
        
        label_width=ancho/30
        bsal_label=tk.Label(self,text="Salchipapas",font=("Brush Script MT",30),bg="red")
        bsal_label.place(x=label_width, y=(1.1*alto/2)-(ual*0.6))
        
        #Boton Hamburguesa
        imagen_bsal=Image.open("boton_hamburguesa.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_salchi= Button(self,text="",image=photo_bsal,command=lambda: parent.show_frame(Hamburguesa))
        button_salchi.image=photo_bsal
        button_salchi.place(x=ancho/5,y=1.1*alto/2)  
        
        label_width=label_width+(ancho/5)
        bsal_label=tk.Label(self,text="Hamburguesas",font=("Brush Script MT",30),bg="red")
        bsal_label.place(x=label_width, y=(1.1*alto/2)-(ual*0.6))
        
        
        #Boton Perro Caliente
        
        imagen_bsal=Image.open("boton_Perrocal.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_salchi= Button(self,text="",image=photo_bsal,command=lambda: parent.show_frame(Perro))
        button_salchi.image=photo_bsal
        button_salchi.place(x=2*(ancho/5),y=1.1*alto/2)
        
        label_width=label_width+(ancho/5)
        bsal_label=tk.Label(self,text="Perros Calientes",font=("Brush Script MT",30),bg="red")
        bsal_label.place(x=label_width-(uan*0.2), y=(1.1*alto/2)-(ual*0.6))
        
        
        #Boton Adicionales
        
        imagen_bsal=Image.open("boton_adicionales.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        button_salchi= Button(self,text="",image=photo_bsal,command=lambda: parent.show_frame(Adicional))
        button_salchi.image=photo_bsal
        button_salchi.place(x=3*(ancho/5),y=1.1*alto/2)
        
        label_width=label_width+(ancho/5)
        bsal_label=tk.Label(self,text="Adicionales",font=("Brush Script MT",30),bg="red")
        bsal_label.place(x=label_width+(uan/10), y=(1.1*alto/2)-(ual*0.6))
        
        #Boton Bebidas
        
        imagen_bsal=Image.open("boton_bebida.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_salchi= Button(self,text="",image=photo_bsal,command=lambda: parent.show_frame(Bebidas))
        button_salchi.image=photo_bsal
        button_salchi.place(x=4*(ancho/5),y=1.1*alto/2)
        
        label_width=label_width+(ancho/5)
        bsal_label=tk.Label(self,text="Bebidas",font=("Brush Script MT",30),bg="red")
        bsal_label.place(x=label_width+(uan*0.3), y=(1.1*alto/2)-(ual*0.6))
        
        #Pregunta
        
        bsal_label=tk.Label(self,text="¿Qué te provoca?",font=("Brush Script MT",40),bg="green")
        bsal_label.place(x=ancho/2.5, y=(1.2*alto/3.3))
        
        
        #Precio Total
        msg=(str(parent.valor)+" $")
        
        Label(self,text = "Valor Total ",font=("Helvetica", 18)).place(x=(ancho/3.13),y=4.213*(alto/5))
        self.cuadro_precio=Entry(self,font=("Helvetica", 18))
        self.cuadro_precio.insert(0,msg)
        self.cuadro_precio.place(x=(ancho/2.40),y=4.205*(alto/5),width=(uan*2.5),height=(ual*0.35))
        
        #Boton Hacer pedido
        button_salchi= Button(self,text="Hacer Pedido",font=("Helvetica", 14),command=lambda: hacer_pedido(self,parent))
        button_salchi.place(x=(ancho/1.65),y=4.2*(alto/5))

class Hamburguesa(tk.Frame):
    def __init__(self,parent):
        
        tk.Frame.__init__(self,parent)
        
        self.cantidadham1=parent.productos["Hamburguesa Sencilla"]
        self.cantidadham2=parent.productos["Hamburguesa Especial"]
        
        def pedido(self,opcion,parent):
            
            if opcion==1:
                self.cantidadham1+=1
                self.ham1.delete(0,END)
                self.ham1.insert(0,str(self.cantidadham1))
                parent.productos["Hamburguesa Sencilla"]+=1
                parent.actualizar_valor()
            elif opcion==2:
                self.cantidadham2+=1
                self.ham2.delete(0,END)
                self.ham2.insert(0,str(self.cantidadham2))
                parent.productos["Hamburguesa Especial"]+=1
                parent.actualizar_valor()
                
        def borrar(self,opcion,parent):
            
            if opcion==1 and self.cantidadham1>0:
                self.cantidadham1-=1
                self.ham1.delete(0,END)
                self.ham1.insert(0,str(self.cantidadham1))
                parent.productos["Hamburguesa Sencilla"]-=1
                parent.actualizar_valor()
                
            elif opcion==2 and self.cantidadham2>0:
                self.cantidadham2-=1
                self.ham2.delete(0,END)
                self.ham2.insert(0,str(self.cantidadham2))
                parent.productos["Hamburguesa Especial"]-=1
                parent.actualizar_valor()
            
        #Background
        background=Image.open("background_vinotinto.jpg")
        background=background.resize((ancho,alto),Image.ANTIALIAS)
        bk_photo = ImageTk.PhotoImage(background,master=self)
        
        bk_label=tk.Label(self,image=bk_photo)
        bk_label.image=bk_photo
        bk_label.place(x=0, y=0, relwidth=1, relheight=1)
        #Boton volver
        button_salchi= Button(self,text="Volver",font=("Arial Black", 15),command=lambda: parent.show_frame(Salchi))
        button_salchi.place(x=(ancho/2.18),y=(4.1*alto/5))
        
        #Titulo
        bsal_label=tk.Label(self,text="Hamburguesas",font=("Brush Script MT",40),bg="Orangered2")
        bsal_label.place(x=ancho/2.5, y=(alto/20))
        
        #hamburguesa1
        imagen_bsal=Image.open("hamburguesa1.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
                
        self.ham1=Entry(self,font=("Arial", 19))
        self.ham1.insert(0,str(self.cantidadham1))
        self.ham1.place(x=(ancho/5),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        button_1= Button(self,text="",image=photo_bsal,command=lambda:pedido(self,1,parent))
        button_1.image=photo_bsal
        button_1.place(x=(ancho/5),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Sencilla - $3000",font=("Brush Script MT",30),bg="green3")
        bsal_label.place(x=(ancho/5)+(alto/50), y=(alto/3.75)-alto/10)
        
        imagen_bsal=Image.open("ingredienteshamburguesa1.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/8)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        tb_image_label= tk.Label(self,image=photo_bsal)
        tb_image_label.image=photo_bsal
        tb_image_label.place(x=(ancho/5),y=(alto/3.75)+(int(alto/4))+(alto/50))
        
        borrar1=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda:borrar(self,1,parent))
        borrar1.place(x=(ancho/5)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
        #hamburguesa2
        imagen_bsal=Image.open("hamburguesa2.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
                
        self.ham2=Entry(self,font=("Arial", 19))
        self.ham2.insert(0,str(self.cantidadham2))
        self.ham2.place(x=(3*ancho/5),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        button_2= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,2,parent))
        button_2.image=photo_bsal
        button_2.place(x=(3*ancho/5),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Especial - $7000",font=("Brush Script MT",30),bg="red3")
        bsal_label.place(x=(3*ancho/5)+(alto/50), y=(alto/3.75)-alto/10)
        
        imagen_bsal=Image.open("ingredienteshamburguesa2.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/8)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        tb_image_label= tk.Label(self,image=photo_bsal)
        tb_image_label.image=photo_bsal
        tb_image_label.place(x=(3*ancho/5),y=(alto/3.75)+(int(alto/4))+(alto/50))
        
        borrar2=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,2,parent))
        borrar2.place(x=(3*ancho/5)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))  
        
        
class Bebidas(tk.Frame):
    def __init__(self,parent):
        
        self.cantidadbeb1=parent.productos["Coca Cola"]
        self.cantidadbeb2=parent.productos["Pepsi"]
        self.cantidadbeb3=parent.productos["Aguila"]
        
        def pedido(self,opcion,parent):
            
            if opcion==1:
                self.cantidadbeb1+=1
                self.beb1.delete(0,END)
                self.beb1.insert(0,str(self.cantidadbeb1))
                parent.productos["Coca Cola"]+=1
                parent.actualizar_valor()
            elif opcion==2:
                self.cantidadbeb2+=1
                self.beb2.delete(0,END)
                self.beb2.insert(0,str(self.cantidadbeb2))
                parent.productos["Pepsi"]+=1
                parent.actualizar_valor()
            elif opcion==3:
                self.cantidadbeb3+=1
                self.beb3.delete(0,END)
                self.beb3.insert(0,str(self.cantidadbeb3))
                parent.productos["Aguila"]+=1
                parent.actualizar_valor()
                
        def borrar(self,opcion,parent):
            
            if opcion==1 and self.cantidadbeb1>0:
                self.cantidadbeb1-=1
                self.beb1.delete(0,END)
                self.beb1.insert(0,str(self.cantidadbeb1))
                parent.productos["Coca Cola"]-=1
                parent.actualizar_valor()
                
            elif opcion==2 and self.cantidadbeb2>0:
                self.cantidadbeb2-=1
                self.beb2.delete(0,END)
                self.beb2.insert(0,str(self.cantidadbeb2))
                parent.productos["Pepsi"]-=1
                parent.actualizar_valor()
                
            elif opcion==3 and self.cantidadbeb3>0:
                self.cantidadbeb3-=1
                self.beb3.delete(0,END)
                self.beb3.insert(0,str(self.cantidadbeb3))
                parent.productos["Aguila"]-=1
                parent.actualizar_valor()
                
        
        tk.Frame.__init__(self,parent)
        #Background
        background=Image.open("background_vinotinto.jpg")
        background=background.resize((ancho,alto),Image.ANTIALIAS)
        bk_photo = ImageTk.PhotoImage(background,master=self)
        
        bk_label=tk.Label(self,image=bk_photo)
        bk_label.image=bk_photo
        bk_label.place(x=0, y=0, relwidth=1, relheight=1)
        
        #Titulo
        bsal_label=tk.Label(self,text="Bebidas",font=("Brush Script MT",40),bg="Orangered2")
        bsal_label.place(x=ancho/2.1, y=(alto/20))
        
        #Boton volver
        button_salchi= Button(self,text="Volver",font=("Arial Black", 15),command=lambda: parent.show_frame(Salchi))
        button_salchi.place(x=(ancho/2.25),y=(4.1*alto/5))
        
        #Bebida1
        imagen_bsal=Image.open("bebida1.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/2.6)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_1= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,1,parent))
        button_1.image=photo_bsal
        button_1.place(x=(ancho/7),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="CocaCola 2.5L - $4000",font=("Brush Script MT",30),bg="green3")
        bsal_label.place(x=(ancho/9)+(alto/50), y=(alto/3.75)-alto/10)
        
        self.beb1=Entry(self,font=("Arial", 19))
        self.beb1.insert(0,str(self.cantidadbeb1))
        self.beb1.place(x=(ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar1=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,1,parent))
        borrar1.place(x=(ancho/7)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
        #Bebida2
        imagen_bsal=Image.open("bebida2.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/2.6)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_2= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,2,parent))
        button_2.image=photo_bsal
        button_2.place(x=(3*ancho/7),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Pepsi 1.25L - $3000",font=("Brush Script MT",30),bg="red3")
        bsal_label.place(x=(3.2*ancho/8)+(alto/50), y=(alto/3.75)-alto/10)
        
        self.beb2=Entry(self,font=("Arial", 19))
        self.beb2.insert(0,str(self.cantidadbeb2))
        self.beb2.place(x=(3*ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar2=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,2,parent))
        borrar2.place(x=(3*ancho/7)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
        #Bebida3
        imagen_bsal=Image.open("bebida3.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/2.6)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_3= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,3,parent))
        button_3.image=photo_bsal
        button_3.place(x=(5*ancho/7),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Aguila Light 2L - $4500",font=("Brush Script MT",30),bg="yellow3")
        bsal_label.place(x=(5.3*ancho/8)+(alto/50), y=(alto/3.75)-alto/10)
        
        self.beb3=Entry(self,font=("Arial", 19))
        self.beb3.insert(0,str(self.cantidadbeb3))
        self.beb3.place(x=(5*ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar3=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,3,parent))
        borrar3.place(x=(5*ancho/7)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
class Adicional(tk.Frame):
    def __init__(self,parent):
        tk.Frame.__init__(self,parent)
        
        self.cantidadad1=parent.productos["Papa Frita"]
        self.cantidadad2=parent.productos["Yuca Frita"]
        self.cantidadad3=parent.productos["Queso"]
        
        def pedido(self,opcion,parent):
            
            if opcion==1:
                self.cantidadad1+=1
                self.ad1.delete(0,END)
                self.ad1.insert(0,str(self.cantidadad1))
                parent.productos["Papa Frita"]+=1
                parent.actualizar_valor()
                
            elif opcion==2:
                self.cantidadad2+=1
                self.ad2.delete(0,END)
                self.ad2.insert(0,str(self.cantidadad2))
                parent.productos["Yuca Frita"]+=1
                parent.actualizar_valor()
                
            elif opcion==3:
                self.cantidadad3+=1
                self.ad3.delete(0,END)
                self.ad3.insert(0,str(self.cantidadad3))
                parent.productos["Queso"]+=1
                parent.actualizar_valor()
                
        def borrar(self,opcion,parent):
            
            if opcion==1 and self.cantidadad1>0:
                self.cantidadad1-=1
                self.ad1.delete(0,END)
                self.ad1.insert(0,str(self.cantidadad1))
                parent.productos["Papa Frita"]-=1
                parent.actualizar_valor()
                
            elif opcion==2 and self.cantidadad2>0:
                self.cantidadad2-=1
                self.ad2.delete(0,END)
                self.ad2.insert(0,str(self.cantidadad2))
                parent.productos["Yuca Frita"]-=1
                parent.actualizar_valor()
                
            elif opcion==3 and self.cantidadad3>0:
                self.cantidadad3-=1
                self.ad3.delete(0,END)
                self.ad3.insert(0,str(self.cantidadad3))
                parent.productos["Queso"]-=1
                parent.actualizar_valor()
                
        #Background
        background=Image.open("background_vinotinto.jpg")
        background=background.resize((ancho,alto),Image.ANTIALIAS)
        bk_photo = ImageTk.PhotoImage(background,master=self)
        
        bk_label=tk.Label(self,image=bk_photo)
        bk_label.image=bk_photo
        bk_label.place(x=0, y=0, relwidth=1, relheight=1)
        
        #Titulo
        bsal_label=tk.Label(self,text="Adicional",font=("Brush Script MT",40),bg="Orangered2")
        bsal_label.place(x=ancho/2.2, y=(alto/20))
        
        #Boton volver
        button_salchi= Button(self,text="Volver",font=("Arial Black", 15),command=lambda: parent.show_frame(Salchi))
        button_salchi.place(x=(ancho/2.25),y=(4.1*alto/5))
        
        #Adicional1
        imagen_bsal=Image.open("adicional1.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/2.6)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_1= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,1,parent))
        button_1.image=photo_bsal
        button_1.place(x=(ancho/7),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Papas Fritas - $3000",font=("Brush Script MT",30),bg="green3")
        bsal_label.place(x=(ancho/9)+(alto/50), y=(alto/3.75)-alto/10)
        
        self.ad1=Entry(self,font=("Arial", 19))
        self.ad1.insert(0,str(self.cantidadad1))
        self.ad1.place(x=(ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar1=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,1,parent))
        borrar1.place(x=(ancho/7)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
        #Adicional2
        imagen_bsal=Image.open("adicional2.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/2.6)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_2= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,2,parent))
        button_2.image=photo_bsal
        button_2.place(x=(3*ancho/7),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Yuca frita - $3000",font=("Brush Script MT",30),bg="red3")
        bsal_label.place(x=(3.4*ancho/8)+(alto/50), y=(alto/3.75)-alto/10)
        
        self.ad2=Entry(self,font=("Arial", 19))
        self.ad2.insert(0,str(self.cantidadad2))
        self.ad2.place(x=(3*ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar2=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,2,parent))
        borrar2.place(x=(3*ancho/7)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
        #Adicional3
        imagen_bsal=Image.open("adicional3.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/2.6)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_3= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,3,parent))
        button_3.image=photo_bsal
        button_3.place(x=(5*ancho/7),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Porción de Queso - $3000",font=("Brush Script MT",30),bg="yellow3")
        bsal_label.place(x=(5.3*ancho/8)+(alto/50), y=(alto/3.75)-alto/10)
        
        self.ad3=Entry(self,font=("Arial", 19))
        self.ad3.insert(0,str(self.cantidadad3))
        self.ad3.place(x=(5*ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar3=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,3,parent))
        borrar3.place(x=(5*ancho/7)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
        
        
class Salchipapa(tk.Frame):
    def __init__(self,parent):
        tk.Frame.__init__(self,parent)
        
        self.productos={"Hamburguesa Sencilla":0,"Hamburguesa Especial":0,"Papa Frita":0,"Yuca Frita":0,
                        "Queso":0,"Salchipapa Sencilla":0,"Salchipapa Especial":0,"Salvajada":0,"Perro Sencillo":0,
                        "Perro Especial":0,"Coca Cola":0,"Pepsi":0,"Aguila":0}
        
        self.cantidadsalchi1=parent.productos["Salchipapa Sencilla"]
        self.cantidadsalchi2=parent.productos["Salchipapa Especial"]
        self.cantidadsalchi3=parent.productos["Salvajada"]
        
        
        def pedido(self,opcion,parent):
            
            if opcion==1:
                self.cantidadsalchi1+=1
                self.salchi1.delete(0,END)
                self.salchi1.insert(0,str(self.cantidadsalchi1))
                parent.productos["Salchipapa Sencilla"]+=1
                parent.actualizar_valor()
                
            elif opcion==2:
                self.cantidadsalchi2+=1
                self.salchi2.delete(0,END)
                self.salchi2.insert(0,str(self.cantidadsalchi2))
                parent.productos["Salchipapa Especial"]+=1
                parent.actualizar_valor()
                
            elif opcion==3:
                self.cantidadsalchi3+=1
                self.salchi3.delete(0,END)
                self.salchi3.insert(0,str(self.cantidadsalchi3))
                parent.productos["Salvajada"]+=1
                parent.actualizar_valor()
                
        def borrar(self,opcion,parent):
            
            if opcion==1 and self.cantidadsalchi1>0:
                self.cantidadsalchi1-=1
                self.salchi1.delete(0,END)
                self.salchi1.insert(0,str(self.cantidadsalchi1))
                parent.productos["Salchipapa Sencilla"]-=1
                parent.actualizar_valor()
                
            elif opcion==2 and self.cantidadsalchi2>0:
                self.cantidadsalchi2-=1
                self.salchi2.delete(0,END)
                self.salchi2.insert(0,str(self.cantidadsalchi2))
                parent.productos["Salchipapa Especial"]-=1
                parent.actualizar_valor()
                
            elif opcion==3 and self.cantidadsalchi3>0:
                self.cantidadsalchi3-=1
                self.salchi3.delete(0,END)
                self.salchi3.insert(0,str(self.cantidadsalchi3))
                parent.productos["Salvajada"]-=1
                parent.actualizar_valor()
        
        #Background#Background
        background=Image.open("background_vinotinto.jpg")
        background=background.resize((ancho,alto),Image.ANTIALIAS)
        bk_photo = ImageTk.PhotoImage(background,master=self)
        
        bk_label=tk.Label(self,image=bk_photo)
        bk_label.image=bk_photo
        bk_label.place(x=0, y=0, relwidth=1, relheight=1)
        
        #Titulo
        bsal_label=tk.Label(self,text="Salchipapas",font=("Brush Script MT",40),bg="Orangered2")
        bsal_label.place(x=ancho/2.4, y=(alto/20))
        
        #Boton volver
        button_salchi= Button(self,text="Volver",font=("Arial Black", 15),command=lambda: parent.show_frame(Salchi))
        button_salchi.place(x=(ancho/2.25),y=(4.1*alto/5))
        
        
        #salchipapa1
        imagen_bsal=Image.open("salchipapa1.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_1= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,1,parent))
        button_1.image=photo_bsal
        button_1.place(x=(ancho/7),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Sencilla - $3000",font=("Brush Script MT",30),bg="green3")
        bsal_label.place(x=(ancho/7)+(alto/50), y=(alto/3.75)-alto/10)
        
        imagen_bsal=Image.open("ingredientessalchi1.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/8)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        tb_image_label= tk.Label(self,image=photo_bsal)
        tb_image_label.image=photo_bsal
        tb_image_label.place(x=(ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50))
        
        self.salchi1=Entry(self,font=("Arial", 19))
        self.salchi1.insert(0,str(self.cantidadsalchi1))
        self.salchi1.place(x=(ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar1=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,1,parent))
        borrar1.place(x=(ancho/7)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
        #salchipapa2
        imagen_bsal=Image.open("salchipapa2.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_2= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,2,parent))
        button_2.image=photo_bsal
        button_2.place(x=(3*ancho/7),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Especial - $6000",font=("Brush Script MT",30),bg="red3")
        bsal_label.place(x=(3*ancho/7)+(alto/50), y=(alto/3.75)-alto/10)
        
        imagen_bsal=Image.open("ingredientessalchi2.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/8)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        tb_image_label= tk.Label(self,image=photo_bsal)
        tb_image_label.image=photo_bsal
        tb_image_label.place(x=(3*ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50))
        
        self.salchi2=Entry(self,font=("Arial", 19))
        self.salchi2.insert(0,str(self.cantidadsalchi2))
        self.salchi2.place(x=(3*ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar2=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,2,parent))
        borrar2.place(x=(3*ancho/7)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
                
        #salchipapa3
        
        imagen_bsal=Image.open("salchipapa3.jpg")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_3= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,3,parent))
        button_3.image=photo_bsal
        button_3.place(x=(5*ancho/7),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Salvajada - $20000",font=("Brush Script MT",30),bg="yellow3")
        bsal_label.place(x=(4.9*ancho/7)+(alto/50), y=(alto/3.75)-alto/10)
        
        imagen_bsal=Image.open("ingredientessalchi3.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/8)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        tb_image_label= tk.Label(self,image=photo_bsal)
        tb_image_label.image=photo_bsal
        tb_image_label.place(x=(5*ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50))
        
        self.salchi3=Entry(self,font=("Arial", 19))
        self.salchi3.insert(0,str(self.cantidadsalchi3))
        self.salchi3.place(x=(5*ancho/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar3=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,3,parent))
        borrar3.place(x=(5*ancho/7)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
class Perro(tk.Frame):
    def __init__(self,parent):
        tk.Frame.__init__(self,parent)

        self.cantidadper1=parent.productos["Perro Sencillo"]
        self.cantidadper2=parent.productos["Perro Especial"]

        
        def pedido(self,opcion,parent):
            
            if opcion==1:
                self.cantidadper1+=1
                self.per1.delete(0,END)
                self.per1.insert(0,str(self.cantidadper1))
                parent.productos["Perro Sencillo"]+=1
                parent.actualizar_valor()
            elif opcion==2:
                self.cantidadper2+=1
                self.per2.delete(0,END)
                self.per2.insert(0,str(self.cantidadper2))
                parent.productos["Perro Especial"]+=1
                parent.actualizar_valor()
                
        def borrar(self,opcion,parent):
            
            if opcion==1 and self.cantidadper1>0:
                self.cantidadper1-=1
                self.per1.delete(0,END)
                self.per1.insert(0,str(self.cantidadper1))
                parent.productos["Perro Sencillo"]-=1
                parent.actualizar_valor()
                
            elif opcion==2 and self.cantidadper2>0:
                self.cantidadper2-=1
                self.per2.delete(0,END)
                self.per2.insert(0,str(self.cantidadper2))
                parent.productos["Perro Especial"]-=1
                parent.actualizar_valor()
                
        #Background
        background=Image.open("background_vinotinto.jpg")
        background=background.resize((ancho,alto),Image.ANTIALIAS)
        bk_photo = ImageTk.PhotoImage(background,master=self)
        
        bk_label=tk.Label(self,image=bk_photo)
        bk_label.image=bk_photo
        bk_label.place(x=0, y=0, relwidth=1, relheight=1)
        
        #Titulo
        bsal_label=tk.Label(self,text="Perros Calientes",font=("Brush Script MT",40),bg="Orangered2")
        bsal_label.place(x=ancho/2.7, y=(alto/20))
        
        #Boton volver
        button_salchi= Button(self,text="Volver",font=("Arial Black", 15),command=lambda: parent.show_frame(Salchi))
        button_salchi.place(x=(ancho/2.25),y=(4.1*alto/5))
        
        
        #hamburguesa1
        imagen_bsal=Image.open("perro1.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_1= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,1,parent))
        button_1.image=photo_bsal
        button_1.place(x=(ancho/5),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Sencillo - $3000",font=("Brush Script MT",30),bg="green3")
        bsal_label.place(x=(ancho/5)+(alto/50), y=(alto/3.75)-alto/10)
        
        imagen_bsal=Image.open("ingredientesperro1.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/8)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        tb_image_label= tk.Label(self,image=photo_bsal)
        tb_image_label.image=photo_bsal
        tb_image_label.place(x=(ancho/5),y=(alto/3.75)+(int(alto/4))+(alto/50))
        
        self.per1=Entry(self,font=("Arial", 19))
        self.per1.insert(0,str(self.cantidadper1))
        self.per1.place(x=(ancho/5),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar1=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,1,parent))
        borrar1.place(x=(ancho/5)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
        #hamburguesa2
        imagen_bsal=Image.open("perro2.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/4)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        button_2= Button(self,text="",image=photo_bsal,command=lambda: pedido(self,2,parent))
        button_2.image=photo_bsal
        button_2.place(x=(3*ancho/5),y=alto/3.75)
        
        bsal_label=tk.Label(self,text="Especial - $6000",font=("Brush Script MT",30),bg="red3")
        bsal_label.place(x=(3*ancho/5)+(alto/50), y=(alto/3.75)-alto/10)
        
        imagen_bsal=Image.open("ingredientesperro2.png")
        imagen_bsal=imagen_bsal.resize((int(ancho/5),int(alto/8)),Image.ANTIALIAS)
        photo_bsal= ImageTk.PhotoImage(imagen_bsal,master=self)
        
        tb_image_label= tk.Label(self,image=photo_bsal)
        tb_image_label.image=photo_bsal
        tb_image_label.place(x=(3*ancho/5),y=(alto/3.75)+(int(alto/4))+(alto/50))
        
        self.per2=Entry(self,font=("Arial", 19))
        self.per2.insert(0,str(self.cantidadper2))
        self.per2.place(x=(3*ancho/5),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/100))
        
        borrar2=Button(self,text="Borrar",font=("Arial Black", 13),command=lambda: borrar(self,2,parent))
        borrar2.place(x=(3*ancho/5)+(alto/7),y=(alto/3.75)+(int(alto/4))+(alto/50)+(alto/8)+(alto/15))
        
        
if __name__== "__main__":
    window=Window()
    window.mainloop()






